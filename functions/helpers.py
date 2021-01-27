import pandas as pd
import dash_core_components as dcc
import dash_bootstrap_components as dbc
import dash_html_components as html
import base64
import io
from openpyxl import load_workbook
import difflib


def build_main_table(main_table_data, clickable):
    def navigate_to(name):
        return dcc.Link(f'{name}', href=f'/{name}')
    main_table_data[clickable] = main_table_data[clickable].map(lambda n: navigate_to(n))
    main_table = html.Div(id='main_table2', children=[
            dbc.Table.from_dataframe(main_table_data, striped=True, bordered=True, hover=True, responsive="sm"),
            ])

    return main_table

def build_page_table(main_table_data, clickable):
    def navigate_to(name):
        return dcc.Link(f'{name}', href=f'/{name}')
    
    main_table_data[clickable] = main_table_data[clickable].map(lambda n: navigate_to(n))
    main_table = html.Div(children=[
            dbc.Table.from_dataframe(main_table_data, striped=True, bordered=True, hover=True, responsive="sm"),
            ])

    return main_table


def build_detail_table(data, pathname):    
    table_container = dbc.Container([
            dbc.Table.from_dataframe(data, striped=True, bordered=True, hover=True, responsive="sm")]
        )

    return table_container


def filter_data(name, detail_data, target):
    filter_detail = detail_data.loc[detail_data[target] == name.replace('/', '')]
    return filter_detail

def parse_contents(contents, filename, date, main_sheet, detail_sheet, clickable, target):
    content_type, content_string = contents.split(',')
    decoded = base64.b64decode(content_string)
    if 'csv' in filename:
        # Assume that the user uploaded a CSV file
        df = pd.read_csv(
            io.StringIO(decoded.decode('utf-8')), skiprows=6)
        return df
    elif 'xls' in filename:
        data = {}
        wb = load_workbook(io.BytesIO(decoded), read_only=True)   # open an Excel file and return a workbook
        if main_sheet in wb.sheetnames:
            main_sheet = pd.read_excel(io.BytesIO(decoded), sheet_name=main_sheet, engine='openpyxl',)
            if clickable not in list(main_sheet.columns):
                similar = difflib.get_close_matches(clickable, list(main_sheet.columns))
                return f'{clickable} does not exist in columns names.!! Did you mean {similar} ?', False
            data['main_data'] = main_sheet.to_dict('records')
        else: 
            return f'{main_sheet} does not exist.!!', False
        if detail_sheet in wb.sheetnames:
            detail_sheet = pd.read_excel(io.BytesIO(decoded), sheet_name=detail_sheet, engine='openpyxl')
            if target not in list(detail_sheet.columns):
                similar = difflib.get_close_matches(target, list(detail_sheet.columns))
                return f'{target} does not exist in columns names.!! Did you mean {similar} ?', False
            data['detail_data'] = detail_sheet.to_dict('records')
        else: 
            return f'{detail_sheet} does not exist.!!', False
        
        return data, True

    else:
        return  None, False


# For now this function read data from excels file... But you can adjust it to query data from your database.
def query_database(main_table='executive_summary', detail_table='tests', clickable_field='name', target_field='name'):
    filePath = 'database/example_data.xlsx'
    main_table_data = pd.read_excel(filePath, sheet_name = main_table, engine='openpyxl',)
    detail_table_data = pd.read_excel(filePath, sheet_name=detail_table, engine='openpyxl')

    # this reshaping process is required: 
    data = {
        'main_data': main_table_data.to_dict('records'),
        'detail_data': detail_table_data.to_dict('records'),
        'clickable_field': clickable_field,
        'target_field': target_field,
        'main_sheet_name': main_table,
        'detail_sheet_name': detail_table
    }

    return data